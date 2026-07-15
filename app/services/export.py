from io import BytesIO
from typing import Sequence

from openpyxl import Workbook

from app.models.document import Document

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from pathlib import Path

import fitz

from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
class ExcelExporter:
    """
    Builds Excel workbooks for document export
    """
    
    def __init__(self, documents: Sequence[Document]):
        self.documents = documents
        
    def build(self) -> BytesIO:
        """
        Build the workbook and return it as an in-memory stream.
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Documents"
        
        self._build_header(worksheet)
        self._populate_rows(worksheet)
        
        worksheet.auto_filter.ref = worksheet.dimensions
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
    
    def _build_header(self, worksheet):
        headers = [
        "ID",
        "Extracted Data",
        "Original Document",
        "Uploaded By",
        "Uploaded At",
        "Status",
        ]

        header_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78",
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )
        
        thin = Side(style="thin", color="000000")

        self.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )
        
        for col, title in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col)

            cell.value = title
            cell.font = header_font
            cell.fill = header_fill

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        worksheet.freeze_panes = "A2"

        worksheet.column_dimensions["A"].width = 10
        worksheet.column_dimensions["B"].width = 45
        worksheet.column_dimensions["C"].width = 35
        worksheet.column_dimensions["D"].width = 20
        worksheet.column_dimensions["E"].width = 20
        worksheet.column_dimensions["F"].width = 15
    
def _populate_rows(self, worksheet):
    row = 2

    for document in self.documents:
        worksheet.cell(row=row, column=1).value = document.id

        # Extracted data
        worksheet.cell(
            row=row,
            column=2,
        ).value = self._format_extracted_data(document.extracted_data)
        
        # Column 3 (Original Document)
        self._insert_document_preview(
            worksheet=worksheet,
            row=row,
            document=document,
        )
        
        # Uploaded By
        uploaded_by = worksheet.cell(
            row=row,
            column=4,
        )

        uploaded_by.value = document.uploader.username
        uploaded_by.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        
        # Uploaded At
        uploaded_at = worksheet.cell(
            row=row,
            column=5,
        )
        
        uploaded_at.value = document.uploaded_at.strftime("%d-%m-%Y %H:%M")
        uploaded_at.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        
        # Status
        status = worksheet.cell(
            row=row,
            column=6,
        )
        
        status.value = document.status.value.capitalize()
        status.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        
        worksheet.cell(
            row=row,
            column=2,
        ).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        
        worksheet.cell(
            row=row,
            column=3,
        ).alignment = Alignment(
            vertical="top",
        )
        
        worksheet.row_dimensions[row].height = 170
        
        row += 1
            
    def _format_extracted_data(self, data: dict | None) -> str:
        """
        Convert OCR JSON into a readable multiline string.
        """
        
        if not data:
            return ""
        
        lines = []
        
        for key, value in data.items():
            pretty_key = key.replace("_", " ").title()
            lines.append(f"{pretty_key}: {value}")
            
        return "\n".join(lines)
    
    def _insert_document_preview(
        self,
        worksheet,
        row: int,
        document: Document,
    ):
        path = Path(document.file_path)
        
        if not path.exists():
            return
        
        suffix = path.suffix.lower()
        
        if suffix == ".pdf":
            self._insert_pdf_preview(
                worksheet,
                row,
                path,
            )
        else:
            self._insert_image(
                worksheet,
                row,
                path,
            )
            
    def _insert_image(
        self,
        worksheet,
        row: int,
        image_path: Path,
    ):
        img = ExcelImage(str(image_path))
        
        img.width = 220
        img.height = 150
        
        worksheet.add_image(
            img,
            f"C{row}",
        )
        
    def _insert_pdf_preview(
        self,
        worksheet,
        row: int,
        pdf_path: Path,
    ):
        """
        Render the first page of a PDF and insert it into Excel.
        """
        try:
            pdf = fitz.open(pdf_path)

            if len(pdf) == 0:
                pdf.close()
                return

            page = pdf.load_page(0)

            matrix = fitz.Matrix(2, 2)

            pix = page.get_pixmap(matrix=matrix)

            image_bytes = pix.tobytes("png")

            pdf.close()

            image = Image.open(BytesIO(image_bytes))

            excel_image = ExcelImage(image)

            excel_image.width = 220
            excel_image.height = 150

            worksheet.add_image(
                excel_image,
                f"C{row}",
            )

        except Exception:
            # Skip preview if rendering fails.
            # The export should still succeed.
            return